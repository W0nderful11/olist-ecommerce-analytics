#!/usr/bin/env python3
"""
Analytics & Visualizations for Olist (Assignment #2)

Функции:
- Подключение к PostgreSQL (та же БД `olist_analytics`).
- Выполнение SQL с 2+ JOIN -> pandas.DataFrame.
- Генерация графиков и сохранение в charts/ (единый стиль).
- Экспорт в Excel с форматированием (exports/).
- Интерактивный график Plotly с time slider.

Запуск примеры:
  python3 analytics.py --make-charts
  python3 analytics.py --time-slider
  python3 analytics.py --export-excel
"""

from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass
from typing import Dict, Tuple

import pandas as pd
import importlib.util
import warnings

try:
    import psycopg2
except Exception:
    print("Требуется psycopg2-binary. Установите: pip install psycopg2-binary", file=sys.stderr)
    raise

# SQLAlchemy для корректной работы pandas.read_sql без предупреждений
try:
    from sqlalchemy import create_engine
except Exception:
    create_engine = None  # не обязателен для работы, но убирает предупреждения

# Глобальные пути вывода
CHARTS_DIR = os.path.join(os.getcwd(), "charts")
EXPORTS_DIR = os.path.join(os.getcwd(), "exports")

# Глобальный engine для pandas.read_sql_query (если доступен SQLAlchemy)
SQL_ENGINE = None


def ensure_dirs() -> None:
    os.makedirs(CHARTS_DIR, exist_ok=True)
    os.makedirs(EXPORTS_DIR, exist_ok=True)


def _prefer_venv_python_if_available() -> None:
    """Если рядом есть .venv и текущий python не из .venv — переключаемся до начала тяжёлой работы.
    Это предотвращает двойной прогон (когда позднее make_time_slider делает re-exec).
    """
    if os.environ.get("OLIST_VENV_SWITCHED") == "1":
        return
    venv_dir = os.path.join(os.getcwd(), ".venv", "bin")
    venv_python = os.path.join(venv_dir, "python")
    venv_python3 = os.path.join(venv_dir, "python3")
    try:
        target = None
        if os.path.exists(venv_python3):
            target = venv_python3
        elif os.path.exists(venv_python):
            target = venv_python
        if target and os.path.realpath(sys.executable) != os.path.realpath(target):
            env = os.environ.copy()
            env["OLIST_VENV_SWITCHED"] = "1"
            os.execve(target, [target] + sys.argv, env)
    except Exception:
        # Тихо продолжаем под текущим интерпретатором
        pass


@dataclass
class DBParams:
    host: str = os.environ.get("PGHOST", "localhost")
    port: str = os.environ.get("PGPORT", "5432")
    dbname: str = os.environ.get("PGDATABASE", "olist_analytics")
    user: str = os.environ.get("PGUSER", "postgres")
    password: str = os.environ.get("PGPASSWORD", "postgres")


def connect_db(p: DBParams):
    return psycopg2.connect(
        host=p.host,
        port=p.port,
        dbname=p.dbname,
        user=p.user,
        password=p.password,
    )


def build_sqlalchemy_engine(p: DBParams):
    """Создать SQLAlchemy engine, если установлен пакет."""
    global SQL_ENGINE
    if create_engine is None:
        return None
    # Пароль может содержать спецсимволы — используем URL‑форму через параметры
    url = f"postgresql+psycopg2://{p.user}:{p.password}@{p.host}:{p.port}/{p.dbname}"
    try:
        SQL_ENGINE = create_engine(url)
    except Exception:
        SQL_ENGINE = None
    return SQL_ENGINE


def run_sql_df(conn, sql: str, params: Tuple = None) -> pd.DataFrame:
    """Выполнить SQL и вернуть DataFrame (сохраняя порядок колонок).
    Предпочтительно используем SQLAlchemy engine (если доступен), чтобы избежать предупреждений pandas.
    """
    if SQL_ENGINE is not None:
        try:
            return pd.read_sql_query(sql, SQL_ENGINE, params=params)
        except Exception as e:
            # Совместимость: pandas<2 не дружит с SQLAlchemy 2.x → тихо уходим на psycopg2
            if "OptionEngine" in str(e) or isinstance(e, AttributeError):
                pass
            else:
                raise
    return pd.read_sql_query(sql, conn, params=params)


def print_report(df: pd.DataFrame, chart_type: str, description: str) -> None:
    rows = len(df)
    print(f"→ {chart_type}: {rows} строк. {description}")


def set_matplotlib_style():
    # Lazy-import через __import__, чтобы линтер не ругался в окружениях без пакетов
    matplotlib = __import__("matplotlib")
    matplotlib.use("Agg")  # headless
    sns = __import__("seaborn")
    sns.set_theme(style="whitegrid")


def save_fig(fig, filename: str, title: str = "") -> str:
    if title:
        fig.suptitle(title)
    out = os.path.join(CHARTS_DIR, filename)
    fig.tight_layout(rect=[0, 0.03, 1, 0.95])
    fig.savefig(out, dpi=150)
    print(f"Сохранён график: {out}")
    return out


 


def _pie_category_revenue(conn) -> pd.DataFrame:
    """Выручка по категориям (>=2 JOIN): order_items + products + translation."""
    sql = """
        SELECT COALESCE(t.product_category_name_english, p.product_category_name) AS category,
               ROUND(SUM(oi.price)::NUMERIC, 2) AS revenue
        FROM olist.order_items oi
        JOIN olist.products p ON p.product_id = oi.product_id
        LEFT JOIN olist.product_category_name_translation t
               ON t.product_category_name = p.product_category_name
        GROUP BY category
        HAVING SUM(oi.price) > 0
        ORDER BY revenue DESC
    """
    df = run_sql_df(conn, sql)
    # Top-N + Other
    top_n = 10
    if len(df) > top_n:
        top = df.iloc[:top_n].copy()
        other_sum = df.iloc[top_n:]["revenue"].sum()
        top.loc[len(top)] = {"category": "Other", "revenue": other_sum}
        return top
    return df


def _bar_gmv_by_state(conn) -> pd.DataFrame:
    sql = """
        SELECT c.customer_state AS state,
               ROUND(SUM(oi.price + oi.freight_value)::NUMERIC, 2) AS gmv
        FROM olist.orders o
        JOIN olist.customers c ON c.customer_id = o.customer_id
        JOIN olist.order_items oi ON oi.order_id = o.order_id
        GROUP BY state
        ORDER BY gmv DESC
    """
    return run_sql_df(conn, sql)


def _barh_top_sellers(conn) -> pd.DataFrame:
    sql = """
        SELECT s.seller_id,
               s.seller_state,
               COUNT(DISTINCT oi.order_id) AS orders,
               ROUND(SUM(oi.price)::NUMERIC, 2) AS revenue
        FROM olist.order_items oi
        JOIN olist.sellers s ON s.seller_id = oi.seller_id
        JOIN olist.orders o ON o.order_id = oi.order_id
        GROUP BY s.seller_id, s.seller_state
        ORDER BY revenue DESC
        LIMIT 15
    """
    return run_sql_df(conn, sql)


def _line_monthly_gmv(conn) -> pd.DataFrame:
    sql = """
        WITH last_month AS (
            SELECT date_trunc('month', MAX(order_purchase_timestamp))::date AS mx
            FROM olist.orders
        ),
        monthly AS (
            SELECT date_trunc('month', o.order_purchase_timestamp)::date AS month,
                   SUM(oi.price + oi.freight_value) AS gmv
            FROM olist.orders o
            JOIN olist.order_items oi ON oi.order_id = o.order_id
            JOIN olist.customers c ON c.customer_id = o.customer_id
            GROUP BY 1
        )
        SELECT m.month::date AS month,
               ROUND(m.gmv::NUMERIC, 2) AS gmv
        FROM monthly m, last_month lm
        WHERE m.month BETWEEN (lm.mx - INTERVAL '11 months') AND lm.mx
        ORDER BY 1
    """
    return run_sql_df(conn, sql)


def _hist_order_totals(conn) -> pd.DataFrame:
    sql = """
        WITH order_totals AS (
            SELECT o.order_id,
                   SUM(oi.price + oi.freight_value) AS order_total
            FROM olist.orders o
            JOIN olist.order_items oi ON oi.order_id = o.order_id
            JOIN olist.customers c ON c.customer_id = o.customer_id
            GROUP BY o.order_id
        )
        SELECT order_id, ROUND(order_total::NUMERIC, 2) AS order_total
        FROM order_totals
        WHERE order_total > 0
        ORDER BY order_total
    """
    return run_sql_df(conn, sql)


def _scatter_price_freight_category_review(conn, sample_limit: int = 4000) -> pd.DataFrame:
    sql = f"""
        SELECT oi.price::float AS price,
               oi.freight_value::float AS freight,
               COALESCE(t.product_category_name_english, p.product_category_name) AS category,
               AVG(r.review_score) OVER (PARTITION BY oi.order_id) AS avg_review_per_order
        FROM olist.order_items oi
        JOIN olist.products p ON p.product_id = oi.product_id
        LEFT JOIN olist.product_category_name_translation t
               ON t.product_category_name = p.product_category_name
        JOIN olist.orders o ON o.order_id = oi.order_id
        LEFT JOIN olist.order_reviews r ON r.order_id = o.order_id
        WHERE oi.price IS NOT NULL AND oi.freight_value IS NOT NULL
        LIMIT {int(sample_limit)}
    """
    return run_sql_df(conn, sql)


def build_all_charts(conn) -> None:
    set_matplotlib_style()
    ensure_dirs()

    import matplotlib.pyplot as plt

    # 1) Pie: revenue share by category
    df1 = _pie_category_revenue(conn)
    fig1, ax1 = plt.subplots(figsize=(8, 6))
    ax1.pie(df1["revenue"], labels=df1["category"], autopct="%1.1f%%", startangle=140)
    ax1.axis('equal')
    save_fig(fig1, "01_pie_category_revenue.png", "Круговая: доля выручки по категориям")
    print_report(df1, "pie", "Доля выручки по категориям (Top10 + Other)")

    # 2) Bar: GMV by customer state
    df2 = _bar_gmv_by_state(conn)
    fig2, ax2 = plt.subplots(figsize=(10, 6))
    ax2.bar(df2["state"], df2["gmv"])
    ax2.set_title("Столбчатая: GMV по штатам клиентов")
    ax2.set_xlabel("Штат")
    ax2.set_ylabel("GMV")
    plt.setp(ax2.get_xticklabels(), rotation=45, ha="right")
    from matplotlib.ticker import StrMethodFormatter
    ax2.yaxis.set_major_formatter(StrMethodFormatter('{x:,.0f}'))
    save_fig(fig2, "02_bar_gmv_by_state.png")
    print_report(df2, "bar", "GMV по штатам клиентов")

    # 3) Horizontal bar: top sellers by revenue
    df3 = _barh_top_sellers(conn)
    fig3, ax3 = plt.subplots(figsize=(10, 7))
    labels3 = [f"{row.seller_state} · {row.seller_id}" for _, row in df3.iterrows()]
    ax3.barh(range(len(df3)), df3["revenue"])
    ax3.set_yticks(range(len(df3)))
    ax3.set_yticklabels(labels3)
    ax3.set_title("Топ продавцов по выручке (state · seller_id)")
    ax3.set_xlabel("Выручка")
    ax3.set_ylabel("Продавец")
    plt.setp(ax3.get_yticklabels(), fontsize=8)
    ax3.invert_yaxis()
    save_fig(fig3, "03_barh_top_sellers.png")
    print_report(df3, "barh", "Топ‑15 продавцов по выручке")

    # 4) Line: monthly GMV 12m
    df4 = _line_monthly_gmv(conn)
    fig4, ax4 = plt.subplots(figsize=(10, 5))
    ax4.plot(df4["month"], df4["gmv"], marker="o")
    ax4.set_title("Линейный график: месячная динамика GMV (12 месяцев)")
    ax4.set_xlabel("Месяц")
    ax4.set_ylabel("GMV")
    plt.setp(ax4.get_xticklabels(), rotation=45, ha="right")
    ax4.yaxis.set_major_formatter(StrMethodFormatter('{x:,.0f}'))
    save_fig(fig4, "04_line_monthly_gmv.png")
    print_report(df4, "line", "Месячная динамика GMV за последние 12 месяцев")

    # 5) Histogram: order totals distribution
    df5 = _hist_order_totals(conn)
    fig5, ax5 = plt.subplots(figsize=(10, 5))
    ax5.hist(df5["order_total"], bins=30, edgecolor="black")
    ax5.set_title("Гистограмма: распределение сумм заказов")
    ax5.set_xlabel("Сумма заказа")
    ax5.set_ylabel("Частота")
    save_fig(fig5, "05_hist_order_totals.png")
    print_report(df5, "histogram", "Распределение сумм заказов")

    # 6) Scatter: price vs freight by category (size by avg review)
    df6 = _scatter_price_freight_category_review(conn)
    fig6, ax6 = plt.subplots(figsize=(9, 6))
    categories = df6["category"].astype(str)
    codes, uniques = pd.factorize(categories)
    scatter = ax6.scatter(df6["price"], df6["freight"], c=codes,
                          s=(df6["avg_review_per_order"].fillna(3) * 15), alpha=0.6, cmap="tab20")
    ax6.set_title("Диаграмма рассеяния: цена vs доставка, цвет — категория, размер — рейтинг")
    ax6.set_xlabel("Цена товара")
    ax6.set_ylabel("Стоимость доставки")
    # Легенда категорий: покажем top-12 по частоте для читаемости
    from matplotlib.lines import Line2D
    from matplotlib import colormaps
    top_cats = categories.value_counts().head(12).index.tolist()
    cmap = colormaps.get('tab20')
    code_by_cat = {cat: i for i, cat in enumerate(uniques)}
    handles = []
    for cat in top_cats:
        color = cmap(code_by_cat.get(cat, 0) % 20)
        handles.append(Line2D([0], [0], marker='o', color='w', label=cat,
                              markerfacecolor=color, markersize=8))
    if handles:
        ax6.legend(handles=handles, title="Категория", loc="upper right", fontsize=8)
    save_fig(fig6, "06_scatter_price_freight.png")
    print_report(df6, "scatter", "Связь цены и доставки по категориям; размер — средний рейтинг заказа")


def _import_plotly_or_reexec():
    """Пытаемся импортировать plotly.express. Если не получается, а рядом есть .venv —
    перезапускаем текущий процесс под .venv/python с теми же аргументами.
    Это устраняет ситуацию, когда shell использует системный python3 без plotly.
    """
    try:
        import plotly.express as px  # noqa: F401
        return True
    except ModuleNotFoundError:
        venv_dir = os.path.join(os.getcwd(), ".venv", "bin")
        venv_python = os.path.join(venv_dir, "python")
        venv_python3 = os.path.join(venv_dir, "python3")
        target = venv_python3 if os.path.exists(venv_python3) else (venv_python if os.path.exists(venv_python) else None)
        if target:
            print("Plotly не найден у текущего python (", sys.executable, ") — перезапускаю под .venv только для тайм‑ползунка…")
            env = os.environ.copy()
            env["OLIST_ONLY_TIME_SLIDER"] = "1"
            env["OLIST_VENV_SWITCHED"] = "1"
            os.execve(target, [target] + sys.argv, env)
        else:
            print("Plotly не установлен. Либо активируйте .venv, либо установите:\n  pip install plotly\nЛибо запустите скрипт через ./.venv/bin/python analytics.py --time-slider")
            return False


def _open_existing_slider_html() -> bool:
    """Если Plotly недоступен, но ранее сохранённый HTML существует — откроем его в браузере.
    Возвращает True, если удалось открыть файл.
    """
    try:
        out_dir = os.path.join(os.getcwd(), "out")
        html_path = os.path.join(out_dir, "plotly_timeslider.html")
        if os.path.exists(html_path):
            import webbrowser
            webbrowser.open(f"file://{html_path}")
            print(f"Plotly не установлен — открыт сохранённый HTML: {html_path}")
            return True
        else:
            print("Сохранённый HTML тайм‑ползунка не найден:", html_path)
            return False
    except Exception as e:
        print("Не удалось открыть сохранённый HTML:", e)
        return False


def make_time_slider(conn) -> None:
    if not _import_plotly_or_reexec():
        # Fallback: если Plotly нет — попробуем открыть уже сохранённый HTML
        _open_existing_slider_html()
        return
    import plotly.express as px
    # Настроим рендерер так, чтобы график открылся в браузере, а также сохраним HTML
    try:
        import plotly.io as pio
        if "browser" not in str(pio.renderers.default):
            pio.renderers.default = "browser"
    except Exception:
        pass
    # Полностью SQL-пайплайн: считаем top-8 категорий по общей выручке, 
    # генерируем помесячный календарь и заполняем отсутствующие месяцы нулями.
    sql = """
        WITH monthly AS (
            SELECT date_trunc('month', o.order_purchase_timestamp)::date AS month,
                   COALESCE(t.product_category_name_english, p.product_category_name) AS category,
                   SUM(oi.price + oi.freight_value) AS gmv
            FROM olist.orders o
            JOIN olist.order_items oi ON oi.order_id = o.order_id
            JOIN olist.products p ON p.product_id = oi.product_id
            LEFT JOIN olist.product_category_name_translation t
                   ON t.product_category_name = p.product_category_name
            WHERE o.order_purchase_timestamp IS NOT NULL
            GROUP BY 1, 2
        ), totals AS (
            SELECT category, SUM(gmv) AS total_gmv
            FROM monthly
            GROUP BY 1
        ), top_categories AS (
            SELECT category
            FROM totals
            ORDER BY total_gmv DESC
            LIMIT 8
        ), bounds AS (
            SELECT date_trunc('month', MIN(o.order_purchase_timestamp))::date AS min_month,
                   date_trunc('month', MAX(o.order_purchase_timestamp))::date AS max_month
            FROM olist.orders o
        ), months AS (
            SELECT generate_series((SELECT min_month FROM bounds),
                                   (SELECT max_month FROM bounds),
                                   INTERVAL '1 month')::date AS month
        )
        SELECT to_char(m.month, 'YYYY-MM') AS month,
               tc.category,
               COALESCE(mm.gmv, 0) AS gmv
        FROM months m
        CROSS JOIN top_categories tc
        LEFT JOIN monthly mm ON mm.month = m.month AND mm.category = tc.category
        ORDER BY m.month, gmv DESC
    """
    dff = run_sql_df(conn, sql)
    # гарантируем числовой тип и адекватную шкалу Y
    dff["gmv"] = pd.to_numeric(dff["gmv"])  # на всякий случай
    fig = px.bar(dff, x="category", y="gmv", color="category",
                 animation_frame="month",
                 title="GMV по категориям c ползунком по месяцам")
    try:
        max_gmv = float(dff["gmv"].max()) if not dff.empty else 1.0
        if max_gmv <= 0:
            max_gmv = 1.0
        fig.update_layout(barmode="group", barnorm=None,
                          yaxis_title="GMV", yaxis_tickformat=",",
                          yaxis_range=[0, max_gmv * 1.1])
    except Exception:
        pass
    # Пытаемся открыть интерактивно и сохранить HTML (однократно)
    only_ts = os.environ.get("OLIST_ONLY_TIME_SLIDER") == "1"
    try:
        fig.show()
    finally:
        out_dir = os.path.join(os.getcwd(), "out")
        os.makedirs(out_dir, exist_ok=True)
        html_path = os.path.join(out_dir, "plotly_timeslider.html")
        try:
            # Если мы уже в режиме только тайм‑ползунка (после re-exec) и файл существует,
            # не дублируем сохранение/сообщения.
            if not (only_ts and os.path.exists(html_path)):
                fig.write_html(html_path, include_plotlyjs="cdn", full_html=True)
                print(f"Сохранён интерактивный HTML: {html_path}")
        except Exception as e:
            print(f"Не удалось сохранить HTML: {e}")


def _timeslider_gmv_by_category_df(conn) -> pd.DataFrame:
    """Данные для тайм‑ползунка: помесячный GMV по top‑8 категориям, с полным календарём месяцев (незаполненные = 0)."""
    sql = """
        WITH monthly AS (
            SELECT date_trunc('month', o.order_purchase_timestamp)::date AS month,
                   COALESCE(t.product_category_name_english, p.product_category_name) AS category,
                   SUM(oi.price + oi.freight_value) AS gmv
            FROM olist.orders o
            JOIN olist.order_items oi ON oi.order_id = o.order_id
            JOIN olist.products p ON p.product_id = oi.product_id
            LEFT JOIN olist.product_category_name_translation t
                   ON t.product_category_name = p.product_category_name
            WHERE o.order_purchase_timestamp IS NOT NULL
            GROUP BY 1, 2
        ), totals AS (
            SELECT category, SUM(gmv) AS total_gmv
            FROM monthly
            GROUP BY 1
        ), top_categories AS (
            SELECT category
            FROM totals
            ORDER BY total_gmv DESC
            LIMIT 8
        ), bounds AS (
            SELECT date_trunc('month', MIN(o.order_purchase_timestamp))::date AS min_month,
                   date_trunc('month', MAX(o.order_purchase_timestamp))::date AS max_month
            FROM olist.orders o
        ), months AS (
            SELECT generate_series((SELECT min_month FROM bounds),
                                   (SELECT max_month FROM bounds),
                                   INTERVAL '1 month')::date AS month
        )
        SELECT to_char(m.month, 'YYYY-MM') AS month,
               tc.category,
               COALESCE(mm.gmv, 0) AS gmv
        FROM months m
        CROSS JOIN top_categories tc
        LEFT JOIN monthly mm ON mm.month = m.month AND mm.category = tc.category
        ORDER BY m.month, gmv DESC
    """
    return run_sql_df(conn, sql)


def export_with_formatting(conn) -> None:
    ensure_dirs()
    # Подготовим данные для всех 6 графиков
    df_cat = _pie_category_revenue(conn)                 # 1. pie
    df_state = _bar_gmv_by_state(conn)                   # 2. bar
    df_sellers = _barh_top_sellers(conn)                 # 3. barh
    df_month = _line_monthly_gmv(conn)                   # 4. line
    if not df_month.empty and pd.api.types.is_datetime64_any_dtype(df_month["month"]):
        df_month = df_month.assign(month=df_month["month"].dt.strftime("%Y-%m"))
    df_orders = _hist_order_totals(conn)                 # 5. histogram (источник — суммы заказов)
    df_scatter = _scatter_price_freight_category_review(conn)  # 6. scatter sample
    df_timeslider = _timeslider_gmv_by_category_df(conn)       # данные для Plotly time‑slider

    data = {
        "01_category_revenue": df_cat,
        "02_gmv_by_state": df_state,
        "03_top_sellers": df_sellers,
        "04_monthly_gmv": df_month,
        "05_order_totals": df_orders,
    "06_scatter_sample": df_scatter,
    "07_timeslider_gmv_by_category": df_timeslider,
    }

    filename = os.path.join(EXPORTS_DIR, "olist_report.xlsx")
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for sheet, df in data.items():
            df.to_excel(writer, sheet_name=sheet, index=False)

        # Применим форматирование через openpyxl
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.utils import get_column_letter
        wb = writer.book
        for sheet in data.keys():
            ws = wb[sheet]
            # Freeze headers
            ws.freeze_panes = "B2"
            # Auto filter
            ws.auto_filter.ref = ws.dimensions
            # Найдём числовые колонки (по первой строке с данными)
            if ws.max_row >= 2:
                header = [c.value for c in ws[1]]
                # определим индексы колонок, которые выглядят числовыми по 2-й строке
                numeric_cols = []
                if ws.max_row >= 2:
                    for idx, cell in enumerate(ws[2], start=1):
                        if isinstance(cell.value, (int, float)):
                            numeric_cols.append(idx)
                # Градиент и условное форматирование min->max
                for col_idx in numeric_cols:
                    col_letter = get_column_letter(col_idx)
                    rng = f"{col_letter}2:{col_letter}{ws.max_row}"
                    rule = ColorScaleRule(start_type="min", start_color="FFAA0000",
                                          mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                                          end_type="max", end_color="FF00AA00")
                    ws.conditional_formatting.add(rng, rule)
            # Авто-ширина колонок по максимальной длине значения
            for col in ws.columns:
                try:
                    max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max(10, max_len + 2), 40)
                except Exception:
                    pass

    total_rows = sum(len(df) for df in data.values())
    print(f"Создан файл {os.path.basename(filename)}, {len(data)} листов, {total_rows} строк")





def build_argparser():
    p = argparse.ArgumentParser(description="Olist Analytics — визуализации и экспорт (Assignment #2)")
    p.add_argument("--host", default=os.environ.get("PGHOST", "localhost"))
    p.add_argument("--port", default=os.environ.get("PGPORT", "5432"))
    p.add_argument("--dbname", default=os.environ.get("PGDATABASE", "olist_analytics"))
    p.add_argument("--user", default=os.environ.get("PGUSER", "postgres"))
    p.add_argument("--password", default=os.environ.get("PGPASSWORD", "postgres"))
    p.add_argument("--make-charts", action="store_true", help="Сгенерировать и сохранить 6 графиков в charts/")
    p.add_argument("--time-slider", action="store_true", help="Показать интерактивный Plotly график с ползунком времени")
    p.add_argument("--export-excel", action="store_true", help="Экспорт выбранных таблиц в Excel с форматированием в exports/")
    return p


def main():
    # Ранний переход на .venv, чтобы всё работало под нужным интерпретатором
    _prefer_venv_python_if_available()
    parser = build_argparser()
    args = parser.parse_args()

    dbp = DBParams(args.host, args.port, args.dbname, args.user, args.password)
    print("Подключение к PostgreSQL:", {k: getattr(dbp, k) for k in ("host", "port", "dbname", "user")})
    try:
        # Заглушим шумные предупреждения
        warnings.filterwarnings("ignore", message="pandas only supports SQLAlchemy connectable", category=UserWarning)
        # инициализируем SQLAlchemy engine (если доступен)
        build_sqlalchemy_engine(dbp)
        with connect_db(dbp) as conn:
            only_ts = os.environ.get("OLIST_ONLY_TIME_SLIDER") == "1"
            if args.make_charts and not only_ts:
                build_all_charts(conn)
            if args.time_slider or only_ts:
                make_time_slider(conn)
            if args.export_excel and not only_ts:
                export_with_formatting(conn)
            if not (args.make_charts or args.time_slider or args.export_excel):
                # Поведение по умолчанию
                if not only_ts:
                    print("Флаги не заданы → по умолчанию: генерируем 6 графиков, Excel-отчёт и интерактивный тайм‑ползунок…")
                    build_all_charts(conn)
                    export_with_formatting(conn)
                # Покажем тайм‑ползунок (и под текущим python, и если re-exec в .venv)
                make_time_slider(conn)
    except Exception as e:
        print("Ошибка выполнения:", e, file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
